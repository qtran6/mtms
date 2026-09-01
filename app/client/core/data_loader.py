"""
data_loader.py — Reads BangGia.xlsx and GiaDacBiet.xlsx using openpyxl (no pandas).

Usage:
    from data_loader import load_products, load_price_groups
"""

import sys
from pathlib import Path

from openpyxl import load_workbook


# ── Path resolution ─────────────────────────────────────────────────────────
def _xlsx_path() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent.parent / "BangGia.xlsx"
    else:
        return Path(__file__).parent.parent.parent / "data" / "BangGia.xlsx"


def _gia_dac_biet_path() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent.parent / "GiaDacBiet.xlsx"
    else:
        return Path(__file__).parent.parent.parent / "data" / "GiaDacBiet.xlsx"


_FILE = _xlsx_path()


# ── Helpers ─────────────────────────────────────────────────────────────────
def _is_blank(v) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def _to_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _sheet_to_grid(file_path: Path, sheet_name: str) -> list[list]:
    """Read a sheet into a 2D list (rows of cell values)."""
    wb = load_workbook(file_path, data_only=True, read_only=True)
    ws = wb[sheet_name]
    grid = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return grid


# ── Load products ───────────────────────────────────────────────────────────
def load_products(file_path: Path = _FILE) -> list[dict]:
    """
    Parse BangGia.xlsx and return [{"brand", "name", "price"}, ...].
    """
    file_path = Path(file_path)
    if file_path.is_dir():
        file_path = file_path / "BangGia.xlsx"

    grid = _sheet_to_grid(file_path, "Tổng Quát")
    if not grid:
        return []

    brand_row = grid[0]
    products = []
    n_cols = max(len(r) for r in grid)

    for col in range(0, n_cols - 1, 2):
        brand = brand_row[col] if col < len(brand_row) else None
        if _is_blank(brand):
            continue
        brand = str(brand).strip()

        for row in range(1, len(grid)):
            data_row = grid[row]
            name  = data_row[col]     if col     < len(data_row) else None
            price = data_row[col + 1] if col + 1 < len(data_row) else None

            if _is_blank(name) or str(name).strip() in ("", "."):
                continue

            price = _to_float(price)
            if price is None:
                continue

            products.append({
                "brand": brand,
                "name":  str(name).strip(),
                "price": price,
            })

    return products


# ── Load special price groups ───────────────────────────────────────────────
def load_price_groups() -> list[dict]:
    """
    Parse GiaDacBiet.xlsx into brand blocks:
        [{"brand", "tiers": [{"label", "prices": {name: price}}]}, ...]
    """
    path = _gia_dac_biet_path()
    if not path.exists():
        return []

    try:
        grid = _sheet_to_grid(path, "Tổng Quát")
    except Exception as e:
        print(f"[price_groups] load failed: {e}")
        return []

    if len(grid) < 2:
        return []

    n_cols = max(len(r) for r in grid)

    def classify(col: int) -> str:
        text, num = 0, 0
        for r in range(1, len(grid)):
            data_row = grid[r]
            if col >= len(data_row):
                continue
            v = data_row[col]
            if _is_blank(v):
                continue
            if _to_float(v) is not None:
                num += 1
            else:
                text += 1
        if text == 0 and num == 0:
            return "empty"
        return "name" if text > num else "price"

    def _fmt_label(v) -> str:
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v).strip()

    kinds = [classify(c) for c in range(n_cols)]

    blocks = []
    current = None
    header_row = grid[0]

    for c in range(n_cols):
        kind = kinds[c]
        header = header_row[c] if c < len(header_row) else None

        if kind == "name":
            if current:
                blocks.append(current)
            current = {
                "brand": str(header).strip() if not _is_blank(header) else "",
                "_name_col": c,
                "tiers": [],
            }
        elif kind == "price" and current is not None:
            label = None if _is_blank(header) else _fmt_label(header)
            prices = {}
            for r in range(1, len(grid)):
                data_row = grid[r]
                name_col = current["_name_col"]
                name  = data_row[name_col] if name_col < len(data_row) else None
                price = data_row[c]        if c        < len(data_row) else None
                if _is_blank(name) or str(name).strip() in ("", "."):
                    continue
                price = _to_float(price)
                if price is None:
                    continue
                prices[str(name).strip()] = price
            current["tiers"].append({"label": label, "prices": prices})

    if current:
        blocks.append(current)

    for b in blocks:
        b.pop("_name_col", None)
    return blocks


if __name__ == "__main__":
    products = load_products()
    print(f"Loaded {len(products)} products")
    for p in products[:20]:
        print(p)